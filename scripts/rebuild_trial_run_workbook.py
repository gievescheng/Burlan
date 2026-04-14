from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "柏連上線試跑紀錄表.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
WARN_FILL = PatternFill("solid", fgColor="FFF4CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(bold=True)
NORMAL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="B7C9D6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row_idx: int) -> None:
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = BORDER


def apply_default_grid(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = NORMAL_ALIGNMENT
            if cell.value not in (None, ""):
                cell.border = BORDER


def build_workbook() -> Workbook:
    wb = Workbook()
    ws_guide = wb.active
    ws_guide.title = "使用說明"
    ws_log = wb.create_sheet("試跑總表")
    ws_cases = wb.create_sheet("建議測試案例")

    guide_rows = [
        ("A1", "柏連上線試跑紀錄表"),
        ("A3", "用途"),
        ("B3", "記錄正式上線前的試跑結果，確認哪些功能正常、哪些地方還要修正。"),
        ("A4", "建議做法"),
        ("B4", "先看建議測試案例，再把實際測試結果寫到試跑總表。"),
        ("A5", "判定"),
        ("B5", "通過 / 需觀察 / 失敗"),
        ("A6", "問題類型"),
        ("B6", "畫面顯示 / 資料來源 / 流程操作 / 權限控管 / 其他"),
        ("A8", "建議試跑順序"),
    ]
    for cell_ref, value in guide_rows:
        ws_guide[cell_ref] = value
    ws_guide["A1"].font = Font(size=16, bold=True)
    for cell_ref in ["A3", "A4", "A5", "A6", "A8"]:
        ws_guide[cell_ref].font = HEADER_FONT

    steps = [
        "1. 文件管理 / 文件庫",
        "2. AI 工作台",
        "3. 稽核計畫",
        "4. 不符合管理",
        "5. 校正管理",
        "6. 設備保養",
        "7. 供應商管理",
        "8. 管理審查資料包",
        "9. 訓練管理",
    ]
    for idx, text in enumerate(steps, start=9):
        ws_guide[f"A{idx}"] = text
    ws_guide.column_dimensions["A"].width = 18
    ws_guide.column_dimensions["B"].width = 78
    for row_idx in range(1, 20):
        ws_guide.row_dimensions[row_idx].height = 24
    apply_default_grid(ws_guide)

    log_headers = [
        "編號",
        "測試日期",
        "測試人",
        "模組",
        "測試項目",
        "操作步驟",
        "預期結果",
        "實際結果",
        "判定",
        "問題類型",
        "問題說明",
        "是否需修正",
        "修正狀態",
        "截圖或檔案路徑",
        "備註",
    ]
    ws_log.append(log_headers)
    style_header_row(ws_log, 1)
    ws_log.freeze_panes = "A2"
    log_widths = [8, 12, 12, 18, 24, 36, 28, 28, 10, 14, 28, 12, 14, 28, 20]
    for idx, width in enumerate(log_widths, start=1):
        ws_log.column_dimensions[get_column_letter(idx)].width = width
    for row_idx in range(2, 22):
        ws_log.row_dimensions[row_idx].height = 44
        for col_idx in range(1, len(log_headers) + 1):
            cell = ws_log.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            cell.alignment = NORMAL_ALIGNMENT
        ws_log.cell(row=row_idx, column=1, value=row_idx - 1).alignment = CENTER_ALIGNMENT

    sample_rows = [
        [
            "範例",
            "2026-03-31",
            "你的名字",
            "稽核計畫",
            "從稽核計畫轉不符合",
            "打開一筆稽核計畫，按轉不符合",
            "會自動帶出不符合草稿",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "這列只是示範，可保留或刪除",
        ],
        [
            "範例",
            "2026-03-31",
            "你的名字",
            "供應商管理",
            "查看柏連供應商名單",
            "打開供應商管理，確認是否看到良器、合億、于正瓶罐",
            "顯示柏連供應商資料",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "先從高優先項目開始測",
        ],
    ]
    for offset, row in enumerate(sample_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_log.cell(row=offset, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = NORMAL_ALIGNMENT
            cell.fill = SUB_FILL

    case_headers = ["排序", "模組", "測試項目", "建議操作", "預期結果", "建議優先度"]
    ws_cases.append(case_headers)
    style_header_row(ws_cases, 1)
    ws_cases.freeze_panes = "A2"
    case_rows = [
        [1, "文件管理", "查看正式文件主清單", "打開文件管理並確認是否看到柏連正式文件主清單", "可看到文件編號、名稱、版本、待人工確認狀態", "高"],
        [2, "文件庫", "搜尋文件", "用文件編號或名稱搜尋一份柏連文件", "可正確找到對應文件", "高"],
        [3, "AI 工作台", "匯入主清單文件", "按匯入柏連主清單全部文件", "可正常匯入，且顯示柏連主清單來源", "高"],
        [4, "稽核計畫", "新增或查看一筆稽核計畫", "打開稽核計畫，確認範圍是否可對應柏連正式文件", "可正常保存，且顯示對應程序文件", "高"],
        [5, "稽核計畫", "轉不符合", "從一筆稽核計畫帶出不符合草稿", "可成功跳轉並自動帶入資料", "高"],
        [6, "不符合管理", "新增與保存不符合", "新增一筆不符合並保存", "資料可保存，重開後仍存在", "高"],
        [7, "校正管理", "查看免校與校正狀態", "確認免校儀器、內校儀器、外校儀器顯示是否正確", "能正確顯示免校 / 內校 / 外校", "高"],
        [8, "校正管理", "更新校正", "對一台非免校儀器記錄校正完成", "可保存校正更新，重開後仍存在", "高"],
        [9, "設備保養", "查看柏連設備清單", "確認設備是否為 BRM-001 到 BRM-008", "可看到柏連設備與保養資料", "高"],
        [10, "設備保養", "記錄保養", "新增一筆設備保養紀錄", "資料可保存，重開後仍存在", "高"],
        [11, "供應商管理", "查看供應商名單", "確認是否看到柏連供應商名單，例如良器、合億、于正瓶罐", "顯示柏連供應商與評鑑結果", "高"],
        [12, "供應商管理", "更新評鑑", "補登或修改一筆供應商評鑑", "資料可保存，且不會跳回舊專案資料", "中"],
        [13, "管理審查", "預檢資料包", "到記錄匯出頁，先檢查管理審查資料", "可看到生產、品質、環境、稽核、不符合的資料筆數", "高"],
        [14, "管理審查", "產出資料包", "直接產生管理審查包", "可成功下載 zip 檔", "高"],
        [15, "訓練管理", "新增與刪除人員", "新增一位人員，再刪除確認是否正常", "可新增、刪除、保存", "中"],
        [16, "主控台 / KPI", "查看統計是否合理", "回到主控台與 KPI 頁面檢查統計", "看到的數量與模組內容大致一致", "中"],
        [17, "重新啟動測試", "確認保存資料仍存在", "關閉系統後重新啟動，再回頭看剛剛保存的資料", "剛剛保存的資料仍存在", "高"],
    ]
    for row in case_rows:
        ws_cases.append(row)
    case_widths = [8, 18, 24, 34, 30, 12]
    for idx, width in enumerate(case_widths, start=1):
        ws_cases.column_dimensions[get_column_letter(idx)].width = width
    for row in ws_cases.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = NORMAL_ALIGNMENT
        priority = row[5].value
        if priority == "高":
            for cell in row:
                cell.fill = WARN_FILL
        elif priority == "中":
            for cell in row:
                cell.fill = SUB_FILL

    for ws in (ws_guide, ws_log, ws_cases):
        ws.sheet_view.showGridLines = True

    return wb


def main() -> None:
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    wb.close()
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
