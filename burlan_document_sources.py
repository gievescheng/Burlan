from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from burlan_paths import BURLAN_MASTER_LIST_PATH


def split_multiline_names(value: str) -> list[str]:
    return [line.strip() for line in str(value or "").splitlines() if line.strip()]


def parse_version_from_filename(name: str) -> str:
    text = str(name or "").strip()
    if not text:
        return ""
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    return match.group(1) if match else ""


def choose_matching_name(names: list[str], preferred_file: str, suffixes: tuple[str, ...]) -> str:
    filtered = [name for name in names if Path(name).suffix.lower() in suffixes]
    if not filtered:
        return ""
    preferred_version = parse_version_from_filename(preferred_file)
    if preferred_file and Path(preferred_file).suffix.lower() in suffixes and preferred_file in filtered:
        return preferred_file
    if preferred_version:
        for name in filtered:
            if parse_version_from_filename(name) == preferred_version:
                return name
    return filtered[0]


def resolve_master_document_path(folder_path: str, preferred_file: str, pdf_names: str, word_names: str) -> str:
    folder = Path(str(folder_path or "").strip())
    candidates = []
    if preferred_file:
        candidates.append(folder / preferred_file)
    for name in split_multiline_names(pdf_names):
        candidates.append(folder / name)
    for name in split_multiline_names(word_names):
        candidates.append(folder / name)

    seen = set()
    for candidate in candidates:
        candidate_str = str(candidate)
        if candidate_str in seen:
            continue
        seen.add(candidate_str)
        if candidate.exists() and candidate.is_file():
            return candidate_str
    return str(candidates[0]) if candidates else ""


def load_burlan_master_documents() -> dict:
    if not BURLAN_MASTER_LIST_PATH.exists():
        return {
            "items": [],
            "source_path": str(BURLAN_MASTER_LIST_PATH),
            "count": 0,
            "pending_review_count": 0,
            "message": "找不到柏連正式文件主清單.xlsx",
        }

    workbook = load_workbook(BURLAN_MASTER_LIST_PATH, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        items = []
        pending_review_count = 0

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
            include_flag = str(data.get("是否納入系統") or "").strip()
            if include_flag not in {"是", "Y", "y", "Yes", "YES", "true", "True"}:
                continue

            review_status = str(data.get("判定狀態") or "").strip()
            if review_status == "待人工確認":
                pending_review_count += 1

            folder_path = str(data.get("實際資料夾位置") or "").strip()
            selected_file = str(data.get("暫定正式檔案") or "").strip()
            pdf_text = str(data.get("找到的 PDF 檔") or "")
            word_text = str(data.get("找到的 Word 檔") or "")
            pdf_names = split_multiline_names(pdf_text)
            word_names = split_multiline_names(word_text)
            resolved_path = resolve_master_document_path(
                folder_path,
                selected_file,
                pdf_text,
                word_text,
            )
            pdf_path = resolve_master_document_path(
                folder_path,
                choose_matching_name(pdf_names, selected_file, (".pdf",)),
                pdf_text,
                "",
            )
            word_path = resolve_master_document_path(
                folder_path,
                choose_matching_name(word_names, selected_file, (".docx", ".doc")),
                "",
                word_text,
            )
            file_type = Path(resolved_path).suffix.lower().lstrip(".") if resolved_path else ""
            items.append(
                {
                    "id": str(data.get("文件編號") or "").strip(),
                    "name": str(data.get("文件名稱") or "").strip(),
                    "department": str(data.get("負責單位") or "").strip(),
                    "version": str(data.get("主清單版次") or "").strip(),
                    "issue_date": str(data.get("主清單發行日期") or "").strip(),
                    "path": resolved_path,
                    "pdf_path": pdf_path,
                    "word_path": word_path,
                    "selected_file": selected_file,
                    "folder_path": folder_path,
                    "file_type": file_type,
                    "review_status": review_status,
                    "review_reason": str(data.get("判定原因") or "").strip(),
                    "source_system": "burlan_official_master",
                }
            )
    finally:
        workbook.close()

    items.sort(key=lambda item: item["id"])
    return {
        "items": items,
        "source_path": str(BURLAN_MASTER_LIST_PATH),
        "count": len(items),
        "pending_review_count": pending_review_count,
        "message": "已載入柏連正式文件主清單",
    }


def normalize_document_title(value: str) -> str:
    return re.sub(r"[\s（）()\-_/]+", "", str(value or "").strip()).lower()


def build_burlan_document_name_lookup(documents: list[dict]) -> dict[str, str]:
    lookup = {}
    for item in documents:
        doc_id = str(item.get("id") or "").strip()
        name = str(item.get("name") or "").strip()
        if not doc_id or not name:
            continue
        lookup.setdefault(normalize_document_title(name), doc_id)
    return lookup
